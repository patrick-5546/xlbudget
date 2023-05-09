"""xlbudget file reading and writing."""

import calendar
from logging import getLogger
from typing import Dict, NamedTuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = getLogger(__name__)

FORMAT_ACCOUNTING = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FORMAT_DATE = "MM/DD/YYYY"

MONTH_NAME_0_IND = calendar.month_name[1:]


class ColumnSpecs(NamedTuple):
    name: str
    format: str
    width: int


COLUMNS = [
    ColumnSpecs(name="Date", format=FORMAT_DATE, width=12),
    ColumnSpecs(name="Description", format="", width=20),
    ColumnSpecs(name="Amount", format=FORMAT_ACCOUNTING, width=12),
]


class TablePosition:
    """The state and bounds of a worksheet table.
    Mangled variables and properties are used to create read-only fields.
    """

    def __init__(self, ref: str) -> None:
        start, end = ref.split(":")

        self.__first_col, self.__header_row = coordinate_from_string(start)
        self.next_row = self.__header_row + 1
        self.__first_col_ind = column_index_from_string(self.__first_col)

        self.__last_col, self.__initial_last_row = coordinate_from_string(end)

    @property
    def first_col(self) -> int:
        return self.__first_col_ind

    @property
    def initial_last_row(self) -> int:
        return self.__initial_last_row

    def __repr__(self) -> str:
        return (
            f"{self.__class__.__name__}(next_row={self.next_row}, "
            f"first_col={self.first_col}, initial_last_row={self.initial_last_row})"
        )

    def get_ref(self) -> str:
        last_row = (
            self.initial_last_row
            if self.initial_last_row == self.next_row
            else self.next_row - 1
        )
        return f"{self.__first_col}{self.__header_row}:{self.__last_col}{last_row}"


def create_year_sheet(wb: Workbook, year: int) -> None:
    """Creates a year sheet, with a table for each month.

    Args:
        wb (openpyxl.workbook.workbook.Workbook): The workbook to create the sheet in.
        year (int): The year.
    """
    index = 0
    for sheet_name in wb.sheetnames:
        sheet_year = int(sheet_name)
        if year < sheet_year:
            break
        elif year == sheet_year:
            raise ValueError(f"Year sheet {year} already exists")

        index += 1
    logger.debug(f"Creating sheet {year} at {index=}")
    ws = wb.create_sheet(str(year), index)
    num_tables = len(MONTH_NAME_0_IND)

    for c_start in range(1, (len(COLUMNS) + 1) * num_tables + 1, len(COLUMNS) + 1):
        month_ind = c_start // (len(COLUMNS) + 1)
        month = MONTH_NAME_0_IND[month_ind]
        table_name = _get_table_name(month, year)
        logger.debug(f"creating {table_name} table")

        # table title
        ws.cell(row=1, column=c_start).value = month
        ws.merge_cells(
            start_row=1,
            start_column=c_start,
            end_row=1,
            end_column=c_start + len(COLUMNS) - 2,
        )

        # table sum
        sum = ws.cell(row=1, column=c_start + len(COLUMNS) - 1)
        sum.value = f"=SUM({table_name}[{COLUMNS[-1].name}])"
        sum.number_format = FORMAT_ACCOUNTING
        logger.debug(f"created sum cell {sum.coordinate}='{sum.value}'")

        # table header and formating
        for i in range(len(COLUMNS)):
            c = c_start + i

            # header
            ws.cell(row=2, column=c).value = COLUMNS[i].name

            # column format
            cell = ws.cell(row=3, column=c)
            if COLUMNS[i].format:
                cell.number_format = COLUMNS[i].format

            # column width
            ws.column_dimensions[get_column_letter(c)].width = COLUMNS[i].width

        # create table
        c_start_ltr = get_column_letter(c_start)
        c_end_ltr = get_column_letter(c_start + len(COLUMNS) - 1)
        ref = f"{c_start_ltr}2:{c_end_ltr}3"
        logger.debug(f"creating table {table_name} with {ref=}")
        tab = Table(displayName=table_name, ref=ref)

        # add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style

        ws.add_table(tab)


def update_xlbudget(wb: Workbook, df: pd.DataFrame):
    """Updates an xlbudget file.

    Args:
        wb (openpyxl.workbook.workbook.Workbook): The xlbudget workbook.
        df (pd.DataFrame): The input file dataframe.
    """
    oldest_date, newest_date = df[df.columns[0]].agg(["min", "max"])
    logger.debug(f"{oldest_date=}, {newest_date=}")

    # create year sheets as needed
    for year in range(oldest_date.year, newest_date.year + 1):
        if str(year) not in wb.sheetnames:
            logger.info(f"Creating {year} sheet")
            create_year_sheet(wb, year)

    # initialize table positions dictionary
    # maps worksheet names to dictionaries that map table names to their position.
    table_pos: Dict[str, Dict[str, TablePosition]] = {}
    for year in range(oldest_date.year, newest_date.year + 1):
        sheet_name = str(year)
        table_pos[sheet_name] = {}

        start_month = oldest_date.month if year == oldest_date.year else 1
        end_month = newest_date.month if year == newest_date.year else 12
        for month in range(start_month, end_month + 1):
            month_name = calendar.month_name[month]
            table_name = _get_table_name(month=month_name, year=sheet_name)
            logger.debug(f"Initializing table {table_name} in sheet {sheet_name}")
            ref = wb[sheet_name].tables[table_name].ref
            table_pos[sheet_name][table_name] = TablePosition(ref)

    # update df with transactions in wb
    logger.debug(f"{df.shape=} before checking existing transactions")
    for sheet_name in table_pos.keys():
        ws = wb[sheet_name]

        for pos in table_pos[sheet_name].values():
            is_populated = bool(ws.cell(row=pos.next_row, column=pos.first_col).value)
            if is_populated:
                for r in range(pos.next_row, pos.initial_last_row + 1):
                    transaction = []
                    for i in range(len(COLUMNS)):
                        c = pos.first_col + i
                        transaction.append(ws.cell(row=r, column=c).value)

                    logger.debug(f"Appending {transaction=} to dataframe")
                    # ignore mypy error and implicitly cast to df.dtypes
                    df.loc[len(df)] = transaction  # type: ignore[call-overload]
    df = df_drop_duplicates(df)
    # re-sort transactions to make the oldest transactions come first
    df = df.sort_values(by=list(df.columns), ascending=True)
    logger.debug(f"{df.shape=} after checking existing transactions")

    # write dataframe to wb
    for row in df.itertuples(index=False):
        logger.debug(f"Writing transaction {row} to workbook")

        # get worksheet and table position
        sheet_name, month_name = str(row.Date.year), calendar.month_name[row.Date.month]
        table_name = _get_table_name(month=month_name, year=sheet_name)
        ws, pos = wb[sheet_name], table_pos[sheet_name][table_name]

        # set date cell
        date_cell = ws.cell(row=pos.next_row, column=pos.first_col)
        date_cell.value = row.Date
        date_cell.number_format = FORMAT_DATE

        # set description cell
        ws.cell(row=pos.next_row, column=pos.first_col + 1).value = row.Description

        # set amount cell
        amount_cell = ws.cell(row=pos.next_row, column=pos.first_col + 2)
        amount_cell.value = row.Amount
        amount_cell.number_format = FORMAT_ACCOUNTING

        pos.next_row += 1

    # update table refs
    for sheet_name in table_pos.keys():
        for table_name, pos in table_pos[sheet_name].items():
            tab = wb[sheet_name].tables[table_name]
            ref = pos.get_ref()
            if ref != tab.ref:
                logger.debug(
                    f"Updating ref of table {tab.name} from {tab.ref} to {ref}"
                )
                tab.ref = pos.get_ref()


def df_drop_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    """Checks for duplicate rows, dropping them in place if any.

    Args:
        df (pd.DataFrame): The original dataframe.

    Returns:
        A[n] `pd.DataFrame` without any duplicate rows.
    """
    duplicated = df.duplicated()
    duplicates = df[duplicated]
    if not duplicates.empty:
        logger.warn(f"Dropping duplicate transactions:\n{duplicates}")
        return df[~duplicated]
    return df


def df_drop_ignores(df: pd.DataFrame, ignore: str) -> pd.DataFrame:
    """Checks for rows that start with `ignore`, dropping them in place if any.

    Args:
        df (pd.DataFrame): The original dataframe.
        ignore (str): The string that begins descriptions to ignore.

    Returns:
        A[n] `pd.DataFrame` without any rows that start with `ignore`.
    """
    ignored = df["Description"].str.startswith(ignore)
    ignores = df[ignored]
    if not ignores.empty:
        logger.warn(f"Dropping ignored transactions:\n{ignores}")
        return df[~ignored]
    return df


def df_drop_na(df: pd.DataFrame) -> pd.DataFrame:
    """Checks for rows that contain only `na` values, dropping them in place if any.

    Args:
        df (pd.DataFrame): The original dataframe.

    Returns:
        A[n] `pd.DataFrame` without any rows that are entirely `na`.
    """
    na = df.isna().all(axis=1)
    nas = df[na]
    if not nas.empty:
        logger.warn(f"Dropping rows that contain only `na` values:\n{nas}")
        return df[~na]
    return df


def _get_table_name(month, year):
    return f"_{month}{year}"
